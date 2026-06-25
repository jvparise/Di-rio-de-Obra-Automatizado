import sys
import json
import os
import shutil
from datetime import datetime
import openpyxl
import win32com.client

MESES = {
    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
    5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
    9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
}

TEMPO_TEXTO = {
    'BOM':           'BOM ( X )                 CHUVOSO (   )            CHUVA INTENSA (   )',
    'CHUVOSO':       'BOM (   )                 CHUVOSO ( X )            CHUVA INTENSA (   )',
    'CHUVA INTENSA': 'BOM (   )                 CHUVOSO (   )            CHUVA INTENSA ( X )',
}

# Célula de quantidade de funcionários (coluna C = próprios, D = terceirizados)
WORKERS_ESQUERDA = {
    'MESTRE DE OBRA': 43, 'ENCARREGADO': 44, 'ALMOXARIFE': 45,
    'ENCARREGADO ADM': 46, 'AUX. ADM': 47, 'TEC. SEGURANCA': 48,
    'ENGENHEIRO': 49, 'ESTAGIARIO': 50, 'AUX. TECNICO': 51,
    'VIGIA': 52, 'SERRALHEIRO': 53,
}

WORKERS_DIREITA = {
    'AJUDANTE COMUM': 43, 'CARPINTEIRO': 44, 'ENCANADOR': 45,
    'AJUDANTE PRATICO': 46, 'ELETRICISTA': 47, 'MONTADOR DE ANDAIME': 48,
    'PEDREIRO': 49, 'OPERADOR DE BETONEIRA': 50,
    'OPERADOR DE RETROESCAVADEIRA': 51, 'ARMADOR': 52,
    'PINTOR': 53, 'GESSEIRO': 54,
}

_config_path = os.path.join(os.path.dirname(__file__), 'config.json')
with open(_config_path, 'r', encoding='utf-8') as _f:
    _config = json.load(_f)

TEMPLATE = _config['template_path']
BASE_OBRAS = _config['base_obras']
EMPRESA = _config.get('empresa', '')
ENGENHEIRO_RESP = _config.get('engenheiro_responsavel', '')


def criar_diario(dados, fotos):
    hoje = datetime.now()
    dia = hoje.day
    mes = hoje.month
    ano = hoje.year
    data_str = hoje.strftime('%d-%m-%Y')
    mes_nome = MESES[mes]
    sheet_name = f'{dia:02d}'

    obra = dados.get('obra', 'Obra').strip()

    pasta_dia = os.path.join(BASE_OBRAS, obra, str(ano), mes_nome, data_str)
    os.makedirs(pasta_dia, exist_ok=True)

    nome_excel = f'Diario Obra {obra}.xlsx'
    caminho_excel = os.path.join(pasta_dia, nome_excel)
    shutil.copy2(TEMPLATE, caminho_excel)

    wb = openpyxl.load_workbook(caminho_excel)

    if sheet_name not in wb.sheetnames:
        resultado = {'sucesso': False, 'erro': f'Aba {sheet_name} não encontrada no template.'}
        print(json.dumps(resultado, ensure_ascii=False))
        return

    ws = wb[sheet_name]

    ws['A7'] = f'OBRA: {obra}'
    ws['I7'] = data_str
    if EMPRESA:
        ws['A9'] = f'EMPRESA EXECUTORA: {EMPRESA}'
    if ENGENHEIRO_RESP:
        ws['A10'] = f'ENGENHEIRO RESPONSÁVEL: {ENGENHEIRO_RESP}'

    descricao = dados.get('descricao', '').strip()
    if descricao:
        linhas = descricao.split('\n')
        for i, linha in enumerate(linhas[:20]):
            row = 12 + i
            if row <= 32:
                ws.cell(row=row, column=1, value=linha.strip())

    tempo = dados.get('tempo', 'BOM').upper().strip()
    ws['A34'] = TEMPO_TEXTO.get(tempo, TEMPO_TEXTO['BOM'])

    workers = dados.get('workers', {})
    for nome, qtd in workers.items():
        nome_upper = nome.upper().strip()
        try:
            qtd_int = int(qtd)
        except (ValueError, TypeError):
            continue

        if nome_upper in WORKERS_ESQUERDA:
            row = WORKERS_ESQUERDA[nome_upper]
            ws.cell(row=row, column=3, value=qtd_int)
        elif nome_upper in WORKERS_DIREITA:
            row = WORKERS_DIREITA[nome_upper]
            ws.cell(row=row, column=8, value=qtd_int)

    wb.save(caminho_excel)

    # Gera PDF da aba do dia
    caminho_pdf = caminho_excel.replace('.xlsx', '.pdf')
    try:
        excel_app = win32com.client.Dispatch('Excel.Application')
        excel_app.Visible = False
        excel_app.DisplayAlerts = False
        wb_com = excel_app.Workbooks.Open(caminho_excel)
        ws_com = wb_com.Worksheets(sheet_name)
        ws_com.ExportAsFixedFormat(0, caminho_pdf)
        wb_com.Close(False)
        excel_app.Quit()
    except Exception as e:
        caminho_pdf = None

    for i, foto_path in enumerate(fotos):
        if os.path.exists(foto_path):
            ext = os.path.splitext(foto_path)[1].lower() or '.jpg'
            nome_foto = f'foto_{i+1:02d}{ext}'
            shutil.copy2(foto_path, os.path.join(pasta_dia, nome_foto))

    resultado = {
        'sucesso': True,
        'pasta': pasta_dia,
        'excel': caminho_excel,
        'pdf': caminho_pdf,
        'fotos': len(fotos)
    }
    print(json.dumps(resultado, ensure_ascii=False))


if __name__ == '__main__':
    if len(sys.argv) > 1:
        with open(sys.argv[1], 'r', encoding='utf-8-sig') as f:
            entrada = json.load(f)
    else:
        entrada = json.loads(sys.stdin.buffer.read().decode('utf-8'))
    criar_diario(entrada.get('dados', {}), entrada.get('fotos', []))
