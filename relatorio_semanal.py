# -*- coding: utf-8 -*-
import sys
import json
import os
from datetime import datetime, timedelta
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

_config_path = os.path.join(os.path.dirname(__file__), 'config.json')
with open(_config_path, 'r', encoding='utf-8') as _f:
    _config = json.load(_f)

BASE_OBRAS = _config['base_obras']

MESES = {
    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
    5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
    9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
}

WORKERS_ESQUERDA = {
    'MESTRE DE OBRA': (43, 'C'), 'ENCARREGADO': (44, 'C'), 'ALMOXARIFE': (45, 'C'),
    'ENCARREGADO ADM': (46, 'C'), 'AUX. ADM': (47, 'C'), 'TEC. SEGURANÇA': (48, 'C'),
    'ENGENHEIRO': (49, 'C'), 'ESTAGIÁRIO': (50, 'C'), 'AUX. TÉCNICO': (51, 'C'),
    'VIGIA': (52, 'C'), 'SERRALHEIRO': (53, 'C'),
}

WORKERS_DIREITA = {
    'AJUDANTE COMUM': (43, 'H'), 'CARPINTEIRO': (44, 'H'), 'ENCANADOR': (45, 'H'),
    'AJUDANTE PRÁTICO': (46, 'H'), 'ELETRICISTA': (47, 'H'), 'MONTADOR DE ANDAIME': (48, 'H'),
    'PEDREIRO': (49, 'H'), 'OP. BETONEIRA': (50, 'H'),
    'OP. RETROESCAVADEIRA': (51, 'H'), 'ARMADOR': (52, 'H'),
    'PINTOR': (53, 'H'), 'GESSEIRO': (54, 'H'),
}


def datas_semana():
    hoje = datetime.now()
    segunda = hoje - timedelta(days=hoje.weekday())
    return [segunda + timedelta(days=i) for i in range(hoje.weekday() + 1)]


def ler_dia(caminho_excel, sheet_name):
    try:
        wb = openpyxl.load_workbook(caminho_excel, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return None
        ws = wb[sheet_name]

        rows = list(ws.iter_rows(min_row=1, max_row=55, max_col=10, values_only=True))
        wb.close()

        obra_raw = str(rows[6][0] or '')
        obra = obra_raw.replace('OBRA:', '').strip()

        atividades = []
        for i in range(11, 32):
            val = rows[i][0] if i < len(rows) else None
            if val and str(val).strip():
                atividades.append(str(val).strip())

        workers = {}
        for nome, (row, col) in {**WORKERS_ESQUERDA, **WORKERS_DIREITA}.items():
            r = row - 1
            c = 2 if col == 'C' else 7
            if r < len(rows):
                val = rows[r][c]
                try:
                    if val and int(val) > 0:
                        workers[nome] = int(val)
                except (ValueError, TypeError):
                    pass

        return {'obra': obra, 'atividades': atividades, 'workers': workers}
    except Exception:
        return None


def gerar_relatorio():
    datas = datas_semana()
    segunda = datas[0]
    hoje = datas[-1]

    relatorio_por_obra = {}

    for data in datas:
        data_str = data.strftime('%d-%m-%Y')
        ano = str(data.year)
        mes_nome = MESES[data.month]
        sheet_name = f'{data.day:02d}'

        if not os.path.isdir(BASE_OBRAS):
            continue

        for obra_nome in os.listdir(BASE_OBRAS):
            pasta_dia = os.path.join(BASE_OBRAS, obra_nome, ano, mes_nome, data_str)
            if not os.path.isdir(pasta_dia):
                continue

            xlsx_files = [f for f in os.listdir(pasta_dia) if f.endswith('.xlsx')]
            if not xlsx_files:
                continue

            caminho_excel = os.path.join(pasta_dia, xlsx_files[0])
            dados = ler_dia(caminho_excel, sheet_name)
            if not dados:
                continue

            if obra_nome not in relatorio_por_obra:
                relatorio_por_obra[obra_nome] = []

            relatorio_por_obra[obra_nome].append({
                'data': data_str,
                'atividades': dados['atividades'],
                'workers': dados['workers']
            })

    if not relatorio_por_obra:
        return '📊 Nenhum diário encontrado nesta semana.'

    linhas = [
        f'📊 *RELATÓRIO SEMANAL*',
        f'Semana: {segunda.strftime("%d/%m")} a {hoje.strftime("%d/%m/%Y")}',
        ''
    ]

    for obra_nome, dias in relatorio_por_obra.items():
        linhas.append(f'🏗️ *{obra_nome}*')
        linhas.append(f'📅 Dias registrados: {len(dias)}')

        # Soma total de workers na semana
        total_workers = {}
        for dia in dias:
            for w, q in dia['workers'].items():
                total_workers[w] = total_workers.get(w, 0) + q

        if total_workers:
            linhas.append('👷 Equipe (total semana):')
            for w, q in total_workers.items():
                linhas.append(f'  - {w}: {q}')

        linhas.append('📋 Atividades:')
        for dia in dias:
            linhas.append(f'  *{dia["data"]}*')
            for at in dia['atividades'][:3]:
                linhas.append(f'  • {at}')
            if len(dia['atividades']) > 3:
                linhas.append(f'  • ... +{len(dia["atividades"]) - 3} atividades')

        linhas.append('')

    return '\n'.join(linhas)


if __name__ == '__main__':
    print(gerar_relatorio())
